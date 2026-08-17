from __future__ import annotations

import argparse
import csv
import json
import pickle
import re
import socket
import struct
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import cv2
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    # JetArm runtime recording does not require the optional Excel index.
    # Keep RGB-D recording available even on a minimal robot installation.
    Workbook = None
    Alignment = Font = PatternFill = None



# ============================================================
# 主要控制設定
# ============================================================
JETARM_HOST = "192.168.51.188"
JETARM_PORT = 9999
DATASET_ROOT = Path(__file__).resolve().parent.parent / "data" / "dataset"
CAPTURE_FPS = 30.0
SAMPLE_INTERVAL = 1  # 1=每幀保存（30 FPS）；5=每 5 幀保存（6 FPS）


TASK_PATTERN = re.compile(r"^task_(\d{4})$")
INDEX_HEADERS = [
    "類別", "Task", "日期", "開始時間", "結束時間", "影片",
    "取樣影格數", "RGB數量", "Depth數量", "照片數量", "狀態",
]


class DatasetManager:
    def __init__(self, dataset_root: Path) -> None:
        self.dataset_root = dataset_root.resolve()
        self.dataset_root.mkdir(parents=True, exist_ok=True)
        self.index_path = self.dataset_root / "dataset_index.xlsx"

    def scan_categories(self) -> list[Path]:
        return sorted(
            [p for p in self.dataset_root.iterdir() if p.is_dir() and not p.name.startswith(".")],
            key=lambda p: p.name.casefold(),
        )

    @staticmethod
    def choose_category(categories: list[Path]) -> Path:
        if not categories:
            raise RuntimeError("dataset 內沒有類別資料夾，請先建立例如 dataset/test。")
        print("\n請選擇本次拍攝類別：")
        for i, category in enumerate(categories, 1):
            print(f"[{i}] {category.name}")
        while True:
            raw = input("輸入編號：").strip()
            try:
                selected = int(raw)
            except ValueError:
                print("請輸入數字編號。")
                continue
            if 1 <= selected <= len(categories):
                return categories[selected - 1]
            print("編號超出範圍。")

    @staticmethod
    def next_task_name(category_dir: Path) -> str:
        used: set[int] = set()
        for child in category_dir.iterdir():
            if child.is_dir() and (m := TASK_PATTERN.fullmatch(child.name)):
                used.add(int(m.group(1)))
        number = 1
        while number in used:
            number += 1
        if number > 9999:
            raise RuntimeError("Task 編號已超過 task_9999。")
        return f"task_{number:04d}"

    @staticmethod
    def create_task(category_dir: Path, task_name: str) -> dict[str, Path]:
        task_dir = category_dir / task_name
        if task_dir.exists():
            raise FileExistsError(f"Task 已存在，為避免覆寫已停止：{task_dir}")
        task_dir.mkdir(parents=True)
        paths = {
            "task": task_dir,
            "video": task_dir / "video",
            "photo": task_dir / "photo",
            "rgb": task_dir / "rgb",
            "depth": task_dir / "depth",
            "annotation": task_dir / "annotation",
            "keyframe": task_dir / "keyframe",
        }
        for key, path in paths.items():
            if key != "task":
                path.mkdir(parents=True, exist_ok=True)
        # 事後挑選的 keyframe 索引；不複製圖片。
        (paths["keyframe"] / "keyframes.json").write_text("[]\n", encoding="utf-8")
        return paths

    def rebuild_dataset_index(self) -> None:
        if Workbook is None:
            return
        rows: list[list[Any]] = []
        for category_dir in self.scan_categories():
            tasks = sorted(
                [p for p in category_dir.iterdir() if p.is_dir() and TASK_PATTERN.fullmatch(p.name)],
                key=lambda p: int(TASK_PATTERN.fullmatch(p.name).group(1)),
            )
            for task_dir in tasks:
                metadata: dict[str, Any] = {}
                metadata_path = task_dir / "metadata.json"
                if metadata_path.exists():
                    try:
                        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
                    except (OSError, json.JSONDecodeError):
                        pass
                video = task_dir / "video" / "raw.mp4"
                video_value = str(video.relative_to(self.dataset_root)).replace("\\", "/") if video.exists() else ""
                rows.append([
                    category_dir.name, task_dir.name, metadata.get("date", ""),
                    metadata.get("start_time", ""), metadata.get("end_time", ""),
                    video_value, metadata.get("sampled_frame_count", 0),
                    metadata.get("rgb_count", 0), metadata.get("depth_count", 0),
                    metadata.get("photo_count", 0), metadata.get("status", "資料夾存在"),
                ])

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dataset Index"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:K{max(1, len(rows) + 1)}"
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        for col, header in enumerate(INDEX_HEADERS, 1):
            cell = sheet.cell(1, col, header)
            cell.fill, cell.font = fill, font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            sheet.append(row)
        for index, width in enumerate([14, 14, 12, 20, 20, 42, 14, 12, 12, 12, 12], 1):
            sheet.column_dimensions[chr(64 + index)].width = width
        workbook.save(self.index_path)


class VisionReceiver:
    HEADER_SIZE = 4

    def __init__(self, host: str, port: int, timeout: float = 10.0) -> None:
        self.host, self.port, self.timeout = host, port, timeout
        self.sock: socket.socket | None = None

    def connect(self) -> None:
        self.close()
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(self.timeout)
        print(f"正在連線 JetArm：{self.host}:{self.port}")
        sock.connect((self.host, self.port))
        sock.settimeout(None)
        self.sock = sock
        print("✅ 已連線 JetArm Vision Bridge")

    def recv_exact(self, size: int) -> bytes:
        if self.sock is None:
            raise ConnectionError("尚未建立連線。")
        data = bytearray()
        while len(data) < size:
            chunk = self.sock.recv(size - len(data))
            if not chunk:
                raise ConnectionError("JetArm 已中斷連線。")
            data.extend(chunk)
        return bytes(data)

    def receive_packet(self) -> dict[str, Any]:
        payload_size = struct.unpack("<L", self.recv_exact(self.HEADER_SIZE))[0]
        if payload_size <= 0 or payload_size > 200_000_000:
            raise ValueError(f"不合理的封包大小：{payload_size}")
        packet = pickle.loads(self.recv_exact(payload_size))
        if not isinstance(packet, dict):
            raise TypeError("收到的封包不是 dict。")
        return packet

    def close(self) -> None:
        if self.sock is not None:
            try:
                self.sock.shutdown(socket.SHUT_RDWR)
            except OSError:
                pass
            try:
                self.sock.close()
            except OSError:
                pass
            self.sock = None


class RecordingSession:
    def __init__(self, manager: DatasetManager, category_dir: Path, sample_interval: int, fps: float) -> None:
        self.manager, self.category_dir = manager, category_dir
        self.sample_interval, self.fps = sample_interval, fps
        self.task_name = DatasetManager.next_task_name(category_dir)
        self.paths: dict[str, Path] | None = None
        self.writer: cv2.VideoWriter | None = None
        self.timestamp_file = None
        self.timestamp_writer: csv.writer | None = None
        self.started = False
        self.finished = False
        self.received_frame_id = self.sampled_frame_count = 0
        self.rgb_count = self.depth_count = self.photo_count = 0
        self.start_dt: datetime | None = None
        self.end_dt: datetime | None = None
        self.last_rgb: np.ndarray | None = None
        self.last_depth: np.ndarray | None = None
        self.last_packet_time: float = 0.0

    @property
    def is_recording(self) -> bool:
        return self.started and not self.finished

    @property
    def task_dir(self) -> Path | None:
        return self.paths["task"] if self.paths else None

    def start(self, frame_size: tuple[int, int]) -> None:
        if self.started:
            return
        self.paths = DatasetManager.create_task(self.category_dir, self.task_name)
        self.start_dt = datetime.now()
        video_path = self.paths["video"] / "raw.mp4"
        self.writer = cv2.VideoWriter(str(video_path), cv2.VideoWriter_fourcc(*"mp4v"), self.fps, frame_size)
        if not self.writer.isOpened():
            raise RuntimeError(f"無法建立影片：{video_path}")
        self.timestamp_file = (self.paths["task"] / "timestamp.csv").open("w", newline="", encoding="utf-8-sig")
        self.timestamp_writer = csv.writer(self.timestamp_file)
        self.timestamp_writer.writerow([
            "received_frame_id", "packet_timestamp", "rgb_timestamp", "depth_timestamp",
            "time_diff_ms", "event", "rgb_file", "depth_file",
        ])
        self.started = True
        self._write_metadata("錄影中")
        self.manager.rebuild_dataset_index()
        print(f"🔴 開始錄影：{self.task_dir}")

    def process_frame(self, rgb: np.ndarray, depth: np.ndarray | None, packet: dict[str, Any]) -> None:
        self.last_rgb, self.last_depth = rgb, depth
        self.last_packet_time = float(packet.get("time", 0.0))
        if not self.is_recording:
            return
        self.received_frame_id += 1
        if self.writer is not None:
            self.writer.write(rgb)
        if self.received_frame_id % self.sample_interval == 0:
            self._save_sample(rgb, depth, packet)

    def _save_sample(self, rgb: np.ndarray, depth: np.ndarray | None, packet: dict[str, Any]) -> None:
        assert self.paths is not None
        stem = f"{self.received_frame_id:06d}"
        rgb_rel, depth_rel = Path("rgb") / f"{stem}.jpg", Path("depth") / f"{stem}.png"
        if not cv2.imwrite(str(self.paths["task"] / rgb_rel), rgb):
            raise OSError(f"RGB 儲存失敗：{rgb_rel}")
        self.rgb_count += 1
        depth_value = ""
        if depth is not None:
            if not cv2.imwrite(str(self.paths["task"] / depth_rel), depth):
                raise OSError(f"Depth 儲存失敗：{depth_rel}")
            self.depth_count += 1
            depth_value = depth_rel.as_posix()
        self.sampled_frame_count += 1
        self._append_timestamp(packet, "sample", rgb_rel.as_posix(), depth_value)

    def save_photo(self, packet: dict[str, Any]) -> None:
        if not self.is_recording or self.paths is None or self.last_rgb is None:
            print("⚠️ 請先按 SPACE 開始錄影，再按 R 拍照。")
            return
        self.photo_count += 1
        stem = f"P_{self.photo_count:04d}_F{self.received_frame_id:06d}"
        photo_rel = Path("photo") / f"{stem}.jpg"
        if not cv2.imwrite(str(self.paths["task"] / photo_rel), self.last_rgb):
            self.photo_count -= 1
            raise OSError(f"照片儲存失敗：{photo_rel}")
        self._append_timestamp(packet, "photo", photo_rel.as_posix(), "")
        self._write_metadata("錄影中")
        print(f"📷 已拍照：{photo_rel}")

    def stop(self, status: str = "完成") -> None:
        if not self.is_recording:
            return
        self.end_dt = datetime.now()
        if self.writer is not None:
            self.writer.release()
            self.writer = None
        if self.timestamp_file is not None:
            self.timestamp_file.flush()
            self.timestamp_file.close()
            self.timestamp_file = None
        self.finished = True
        self._write_metadata(status)
        self.manager.rebuild_dataset_index()
        print(f"⏹️ 錄影結束：{self.task_dir}")

    def _append_timestamp(self, packet: dict[str, Any], event: str, rgb_file: str, depth_file: str) -> None:
        if self.timestamp_writer is None:
            return
        self.timestamp_writer.writerow([
            self.received_frame_id,
            f"{float(packet.get('time', 0.0)):.9f}",
            f"{float(packet.get('rgb_timestamp', 0.0)):.9f}",
            f"{float(packet.get('depth_timestamp', 0.0)):.9f}",
            f"{float(packet.get('time_diff_ms', 0.0)):.3f}",
            event, rgb_file, depth_file,
        ])
        if self.timestamp_file is not None:
            self.timestamp_file.flush()

    def _write_metadata(self, status: str) -> None:
        if self.paths is None:
            return
        metadata = {
            "category": self.category_dir.name, "task": self.task_name,
            "date": self.start_dt.strftime("%Y-%m-%d") if self.start_dt else "",
            "start_time": self.start_dt.strftime("%Y-%m-%d %H:%M:%S") if self.start_dt else "",
            "end_time": self.end_dt.strftime("%Y-%m-%d %H:%M:%S") if self.end_dt else "",
            "status": status, "sample_interval": self.sample_interval,
            "recording_fps": self.fps, "received_frame_count": self.received_frame_id,
            "sampled_frame_count": self.sampled_frame_count, "rgb_count": self.rgb_count,
            "depth_count": self.depth_count, "photo_count": self.photo_count,
            "video_file": "video/raw.mp4", "photo_dir": "photo",
            "keyframe_index": "keyframe/keyframes.json",
        }
        (self.paths["task"] / "metadata.json").write_text(
            json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8"
        )


def decode_packet(packet: dict[str, Any]) -> tuple[np.ndarray, np.ndarray | None]:
    rgb_bytes = packet.get("rgb")
    if not isinstance(rgb_bytes, (bytes, bytearray)):
        raise ValueError("封包缺少 RGB JPEG bytes。")
    rgb = cv2.imdecode(np.frombuffer(rgb_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
    if rgb is None:
        raise ValueError("RGB JPEG 解碼失敗。")
    depth = packet.get("depth")
    if depth is not None and not isinstance(depth, np.ndarray):
        depth = np.asarray(depth)
    return rgb, depth


def draw_overlay(frame: np.ndarray, category: str, session: RecordingSession) -> np.ndarray:
    display = frame.copy()
    state = "REC" if session.is_recording else "PREVIEW"
    lines = [
        f"Category: {category}", f"Next/Current Task: {session.task_name}", f"State: {state}",
        "SPACE: start/stop task | R: photo | Q/ESC: quit",
    ]
    for i, text in enumerate(lines):
        cv2.putText(display, text, (12, 28 + i * 28), cv2.FONT_HERSHEY_SIMPLEX,
                    0.65, (255, 255, 255), 2, cv2.LINE_AA)
    return display


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="JetArm 智慧焊接資料收集程式")
    parser.add_argument("--host", default=JETARM_HOST, help="JetArm IP")
    parser.add_argument("--port", type=int, default=JETARM_PORT, help="TCP Port")
    parser.add_argument(
        "--dataset-root", type=Path,
        default=DATASET_ROOT,
        help="資料集根目錄，預設為 JetArm_Project/data/dataset",
    )
    parser.add_argument("--sample-interval", type=int, default=SAMPLE_INTERVAL, help="每隔幾幀保存一組 RGB/Depth")
    parser.add_argument("--fps", type=float, default=CAPTURE_FPS, help="輸出影片 FPS")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.sample_interval <= 0 or args.fps <= 0:
        raise SystemExit("--sample-interval 與 --fps 必須大於 0。")
    manager = DatasetManager(args.dataset_root)
    manager.rebuild_dataset_index()
    try:
        category_dir = manager.choose_category(manager.scan_categories())
    except RuntimeError as exc:
        print(f"❌ {exc}\n目前 dataset 路徑：{manager.dataset_root}")
        return 1

    receiver = VisionReceiver(args.host, args.port)
    session = RecordingSession(manager, category_dir, args.sample_interval, args.fps)
    latest_packet: dict[str, Any] = {}
    print(f"\n已選擇類別：{category_dir.name}")
    print("SPACE 開始/停止每一個 Task；停止後再按 SPACE 會建立下一個 Task。")

    try:
        receiver.connect()
        while True:
            latest_packet = receiver.receive_packet()
            rgb, depth = decode_packet(latest_packet)
            session.process_frame(rgb, depth, latest_packet)
            cv2.imshow("JetArm Dataset Recorder", draw_overlay(rgb, category_dir.name, session))
            key = cv2.waitKey(1) & 0xFF

            if key == ord(" "):
                if not session.is_recording:
                    # 已完成上一個 Task 時，建立全新的 session，畫面不退出。
                    if session.started:
                        session = RecordingSession(manager, category_dir, args.sample_interval, args.fps)
                    h, w = rgb.shape[:2]
                    session.start((w, h))
                else:
                    session.stop("完成")
                    print(f"下一次按 SPACE 將建立：{DatasetManager.next_task_name(category_dir)}")
            elif key in (ord("r"), ord("R")):
                session.save_photo(latest_packet)
            elif key in (27, ord("q"), ord("Q")):
                if session.is_recording:
                    session.stop("中斷")
                break

    except KeyboardInterrupt:
        print("\n收到 Ctrl+C。")
        if session.is_recording:
            session.stop("中斷")
    except (ConnectionError, OSError, ValueError, TypeError, pickle.UnpicklingError, RuntimeError) as exc:
        print(f"❌ 執行錯誤：{exc}")
        if session.is_recording:
            session.stop("中斷")
        return 1
    finally:
        receiver.close()
        cv2.destroyAllWindows()
        manager.rebuild_dataset_index()
    return 0


if __name__ == "__main__":
    sys.exit(main())
