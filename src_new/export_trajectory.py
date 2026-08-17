from __future__ import annotations

import argparse
import csv
import json
import math
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import cv2
import numpy as np


# ============================================================
# 預設設定
# ============================================================

DEPTH_WINDOW = 11
MIN_VALID_DEPTH_PIXELS = 12

MIN_DEPTH_MM = 150
MAX_DEPTH_MM = 3000

TEMPORAL_WINDOW = 5
STATE_WINDOW = 3

MAX_INTERPOLATE_GAP = 3
MAX_CAMERA_STEP_MM = 100.0

# RGB 相機內參，對應 640 × 480
FX = 456.0071
FY = 456.0071
CX = 325.24936
CY = 240.54808

INTRINSIC_WIDTH = 640
INTRINSIC_HEIGHT = 480


TRAJECTORY_FIELDS = [
    "source_frame_id",
    "sample_id",
    "timestamp_s",
    "image_name",
    "source",
    "depth_path",

    "state_raw",
    "state_filtered",
    "confidence",

    "x_px_raw",
    "y_px_raw",
    "x_px_filtered",
    "y_px_filtered",

    "image_width",
    "image_height",
    "depth_width",
    "depth_height",
    "depth_x_px",
    "depth_y_px",

    "depth_exact_mm",
    "depth_local_median_mm",
    "depth_temporal_median_mm",
    "depth_valid_pixels",
    "depth_valid_ratio",
    "depth_mapping",
    "depth_source",

    "camera_x_mm",
    "camera_y_mm",
    "camera_z_mm",
    "step_distance_mm",

    "interpolated",
    "outlier",
    "valid",
    "invalid_reason",
    "quality",

    "segment_type",
    "segment_id",

    "coordinate_frame",
    "robot_execution_ready",
]


# ============================================================
# 參數
# ============================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "將 infer_yolo.py 的 predictions.csv "
            "轉換成深度與相機 XYZ 軌跡"
        )
    )

    parser.add_argument(
        "--predictions",
        type=Path,
        required=True,
        help="infer_yolo.py 輸出的 predictions.csv",
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="未指定時輸出到 predictions.csv 同層的 trajectory",
    )

    parser.add_argument(
        "--depth-window",
        type=int,
        default=DEPTH_WINDOW,
    )

    parser.add_argument(
        "--min-valid-depth-pixels",
        type=int,
        default=MIN_VALID_DEPTH_PIXELS,
    )

    parser.add_argument(
        "--min-depth-mm",
        type=int,
        default=MIN_DEPTH_MM,
    )

    parser.add_argument(
        "--max-depth-mm",
        type=int,
        default=MAX_DEPTH_MM,
    )

    parser.add_argument(
        "--temporal-window",
        type=int,
        default=TEMPORAL_WINDOW,
    )

    parser.add_argument(
        "--state-window",
        type=int,
        default=STATE_WINDOW,
    )

    parser.add_argument(
        "--max-interpolate-gap",
        type=int,
        default=MAX_INTERPOLATE_GAP,
    )

    parser.add_argument(
        "--max-camera-step-mm",
        type=float,
        default=MAX_CAMERA_STEP_MM,
    )

    parser.add_argument("--fx", type=float, default=FX)
    parser.add_argument("--fy", type=float, default=FY)
    parser.add_argument("--cx", type=float, default=CX)
    parser.add_argument("--cy", type=float, default=CY)

    parser.add_argument(
        "--intrinsic-width",
        type=int,
        default=INTRINSIC_WIDTH,
    )

    parser.add_argument(
        "--intrinsic-height",
        type=int,
        default=INTRINSIC_HEIGHT,
    )

    args = parser.parse_args()

    for name in (
        "depth_window",
        "temporal_window",
        "state_window",
    ):
        value = getattr(args, name)

        if value <= 0 or value % 2 == 0:
            parser.error(
                f"--{name.replace('_', '-')} "
                "必須是大於 0 的奇數"
            )

    return args


# ============================================================
# 基本工具
# ============================================================

def project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def resolve_path(path: Path) -> Path:
    if path.is_absolute():
        return path.expanduser().resolve()

    return (
        project_root()
        / path
    ).expanduser().resolve()


def to_float(value: Any) -> float | None:
    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def to_int(value: Any) -> int | None:
    number = to_float(value)

    if number is None:
        return None

    return int(round(number))


def is_number(value: Any) -> bool:
    return (
        value is not None
        and isinstance(value, (int, float, np.number))
        and math.isfinite(float(value))
    )


def read_predictions(
    path: Path,
) -> list[dict[str, Any]]:
    with path.open(
        "r",
        newline="",
        encoding="utf-8-sig",
    ) as file:
        reader = csv.DictReader(file)
        source_rows = list(reader)

    rows: list[dict[str, Any]] = []
    manifest_sources = load_manifest_sources()

    for index, source_row in enumerate(source_rows):
        source_path = Path(
            str(source_row["source"])
        )

        if not source_path.is_absolute():
            source_path = (
                project_root()
                / source_path
            ).resolve()

        # infer_yolo.py may run on the flattened YOLO dataset.  Those images
        # live under images/{train,val,test} and do not have a sibling depth
        # directory.  Resolve them back to the original task RGB image through
        # manifest.csv so depth lookup remains valid.
        original_source = manifest_sources.get(source_path.name)

        if original_source is not None:
            source_path = original_source

        row = {
            "source_frame_id": (
                to_int(
                    source_row.get(
                        "source_frame_id"
                    )
                )
                or index
            ),
            "sample_id": (
                to_int(
                    source_row.get("sample_id")
                )
                or index
            ),
            "timestamp_s": to_float(
                source_row.get("timestamp_s")
            ),
            "image_name": source_path.name,
            "source": str(source_path),

            "state_raw": str(
                source_row.get(
                    "state",
                    "NONE",
                )
            ),
            "confidence": (
                to_float(
                    source_row.get(
                        "confidence"
                    )
                )
                or 0.0
            ),

            "x_px_raw": to_float(
                source_row.get("x_px")
            ),
            "y_px_raw": to_float(
                source_row.get("y_px_top")
            ),

            "image_width": (
                to_int(
                    source_row.get(
                        "image_width"
                    )
                )
                or 0
            ),
            "image_height": (
                to_int(
                    source_row.get(
                        "image_height"
                    )
                )
                or 0
            ),
        }

        rows.append(row)

    return rows


def load_manifest_sources() -> dict[str, Path]:
    manifest_path = (
        project_root()
        / "data"
        / "yolo_tip_state_dataset"
        / "manifest.csv"
    )

    if not manifest_path.is_file():
        return {}

    sources: dict[str, Path] = {}

    with manifest_path.open(
        "r",
        newline="",
        encoding="utf-8-sig",
    ) as file:
        reader = csv.DictReader(file)

        for row in reader:
            image_name = Path(
                str(row.get("output_image", ""))
            ).name

            source_text = str(
                row.get("source_image", "")
            ).strip()

            if not image_name or not source_text:
                continue

            source_path = Path(source_text)

            if not source_path.is_absolute():
                source_path = (
                    project_root()
                    / source_path
                ).resolve()

            sources[image_name] = source_path

    return sources


# ============================================================
# Depth 配對與空間中位數
# ============================================================

def matching_depth_path(
    source_path: Path,
) -> Path:
    rgb_dir = source_path.parent
    task_dir = rgb_dir.parent

    return (
        task_dir
        / "depth"
        / f"{source_path.stem}.png"
    )


def load_depth(
    path: Path,
) -> np.ndarray | None:
    if not path.is_file():
        return None

    depth = cv2.imread(
        str(path),
        cv2.IMREAD_UNCHANGED,
    )

    if depth is None:
        return None

    if depth.ndim == 3:
        depth = depth[:, :, 0]

    return depth


def map_rgb_to_depth(
    x: float,
    y: float,
    rgb_width: int,
    rgb_height: int,
    depth_width: int,
    depth_height: int,
) -> tuple[int, int, str]:
    if (
        rgb_width == depth_width
        and rgb_height == depth_height
    ):
        depth_x = int(round(x))
        depth_y = int(round(y))

        mapping = "same_size"

    else:
        depth_x = int(
            round(
                x
                * max(depth_width - 1, 1)
                / max(rgb_width - 1, 1)
            )
        )

        depth_y = int(
            round(
                y
                * max(depth_height - 1, 1)
                / max(rgb_height - 1, 1)
            )
        )

        mapping = "scaled_pixel_approx"

    depth_x = int(
        np.clip(
            depth_x,
            0,
            depth_width - 1,
        )
    )

    depth_y = int(
        np.clip(
            depth_y,
            0,
            depth_height - 1,
        )
    )

    return depth_x, depth_y, mapping


def robust_depth(
    depth: np.ndarray | None,
    x: float | None,
    y: float | None,
    rgb_width: int,
    rgb_height: int,
    args: argparse.Namespace,
) -> dict[str, Any]:
    result = {
        "depth_width": None,
        "depth_height": None,
        "depth_x_px": None,
        "depth_y_px": None,
        "depth_exact_mm": None,
        "depth_local_median_mm": None,
        "depth_valid_pixels": 0,
        "depth_valid_ratio": 0.0,
        "depth_mapping": "unavailable",
    }

    if (
        depth is None
        or x is None
        or y is None
    ):
        return result

    depth_height, depth_width = depth.shape[:2]

    depth_x, depth_y, mapping = (
        map_rgb_to_depth(
            x=x,
            y=y,
            rgb_width=rgb_width,
            rgb_height=rgb_height,
            depth_width=depth_width,
            depth_height=depth_height,
        )
    )

    result.update({
        "depth_width": depth_width,
        "depth_height": depth_height,
        "depth_x_px": depth_x,
        "depth_y_px": depth_y,
        "depth_mapping": mapping,
    })

    exact = int(depth[depth_y, depth_x])

    if (
        args.min_depth_mm
        <= exact
        <= args.max_depth_mm
    ):
        result["depth_exact_mm"] = exact

    radius = args.depth_window // 2

    x0 = max(0, depth_x - radius)
    x1 = min(
        depth_width,
        depth_x + radius + 1,
    )

    y0 = max(0, depth_y - radius)
    y1 = min(
        depth_height,
        depth_y + radius + 1,
    )

    region = depth[y0:y1, x0:x1].astype(
        np.float64
    )

    valid = region[
        (region >= args.min_depth_mm)
        & (region <= args.max_depth_mm)
    ]

    if valid.size == 0:
        return result

    initial_median = float(
        np.median(valid)
    )

    deviation = np.abs(
        valid - initial_median
    )

    mad = float(
        np.median(deviation)
    )

    tolerance = max(
        20.0,
        3.0 * 1.4826 * mad,
    )

    filtered = valid[
        deviation <= tolerance
    ]

    valid_count = int(filtered.size)

    result["depth_valid_pixels"] = (
        valid_count
    )

    result["depth_valid_ratio"] = round(
        valid_count / region.size,
        6,
    )

    if (
        valid_count
        >= args.min_valid_depth_pixels
    ):
        result[
            "depth_local_median_mm"
        ] = round(
            float(np.median(filtered)),
            3,
        )

    return result


def attach_depth(
    rows: list[dict[str, Any]],
    args: argparse.Namespace,
) -> None:
    for index, row in enumerate(rows):
        source_path = Path(row["source"])

        depth_path = matching_depth_path(
            source_path
        )

        row["depth_path"] = str(depth_path)

        depth = load_depth(depth_path)

        result = robust_depth(
            depth=depth,
            x=row["x_px_raw"],
            y=row["y_px_raw"],
            rgb_width=row["image_width"],
            rgb_height=row["image_height"],
            args=args,
        )

        row.update(result)

        if row["timestamp_s"] is None:
            row["timestamp_s"] = round(
                index / 6.0,
                6,
            )


# ============================================================
# 時間濾波
# ============================================================

def centered_median(
    rows: list[dict[str, Any]],
    index: int,
    field: str,
    window: int,
) -> float | None:
    radius = window // 2

    start = max(0, index - radius)
    end = min(
        len(rows),
        index + radius + 1,
    )

    values = [
        float(row[field])
        for row in rows[start:end]
        if is_number(row.get(field))
    ]

    if not values:
        return None

    return round(
        float(np.median(values)),
        3,
    )


def filtered_states(
    rows: list[dict[str, Any]],
    window: int,
) -> list[str]:
    radius = window // 2
    output: list[str] = []

    for index, row in enumerate(rows):
        start = max(0, index - radius)
        end = min(
            len(rows),
            index + radius + 1,
        )

        states = [
            item["state_raw"]
            for item in rows[start:end]
        ]

        counts = Counter(states)
        highest = max(counts.values())

        candidates = {
            state
            for state, count in counts.items()
            if count == highest
        }

        original = row["state_raw"]

        if original in candidates:
            output.append(original)
        elif "DOWN" in candidates:
            output.append("DOWN")
        elif "UP" in candidates:
            output.append("UP")
        else:
            output.append("NONE")

    return output


def temporal_filter(
    rows: list[dict[str, Any]],
    args: argparse.Namespace,
) -> None:
    states = filtered_states(
        rows,
        args.state_window,
    )

    for index, row in enumerate(rows):
        row["state_filtered"] = states[index]

        row["x_px_filtered"] = (
            centered_median(
                rows,
                index,
                "x_px_raw",
                args.temporal_window,
            )
        )

        row["y_px_filtered"] = (
            centered_median(
                rows,
                index,
                "y_px_raw",
                args.temporal_window,
            )
        )

        row[
            "depth_temporal_median_mm"
        ] = centered_median(
            rows,
            index,
            "depth_local_median_mm",
            args.temporal_window,
        )

        if (
            row[
                "depth_temporal_median_mm"
            ]
            is not None
        ):
            row["depth_source"] = (
                "temporal_local_median"
            )
        else:
            row["depth_source"] = "invalid"

        row["interpolated"] = False


# ============================================================
# 1～3 幀缺失內插
# ============================================================

def interpolate_field(
    rows: list[dict[str, Any]],
    field: str,
    start: int,
    end: int,
) -> bool:
    previous_value = rows[
        start - 1
    ].get(field)

    next_value = rows[
        end + 1
    ].get(field)

    if (
        not is_number(previous_value)
        or not is_number(next_value)
    ):
        return False

    gap = end - start + 1

    for offset, index in enumerate(
        range(start, end + 1),
        start=1,
    ):
        ratio = offset / (gap + 1)

        rows[index][field] = round(
            float(previous_value)
            + (
                float(next_value)
                - float(previous_value)
            )
            * ratio,
            3,
        )

    return True


def interpolate_short_gaps(
    rows: list[dict[str, Any]],
    args: argparse.Namespace,
) -> None:
    index = 0

    while index < len(rows):
        row = rows[index]

        missing = (
            row["state_filtered"] != "NONE"
            and (
                row["x_px_filtered"] is None
                or row["y_px_filtered"] is None
                or row[
                    "depth_temporal_median_mm"
                ] is None
            )
        )

        if not missing:
            index += 1
            continue

        start = index

        while (
            index < len(rows)
            and rows[index][
                "state_filtered"
            ] != "NONE"
            and (
                rows[index][
                    "x_px_filtered"
                ] is None
                or rows[index][
                    "y_px_filtered"
                ] is None
                or rows[index][
                    "depth_temporal_median_mm"
                ] is None
            )
        ):
            index += 1

        end = index - 1
        gap = end - start + 1

        bounded = (
            start > 0
            and end + 1 < len(rows)
        )

        if (
            not bounded
            or gap
            > args.max_interpolate_gap
        ):
            continue

        success = True

        for field in (
            "x_px_filtered",
            "y_px_filtered",
            "depth_temporal_median_mm",
        ):
            success = (
                interpolate_field(
                    rows,
                    field,
                    start,
                    end,
                )
                and success
            )

        if success:
            for fill_index in range(
                start,
                end + 1,
            ):
                rows[fill_index][
                    "interpolated"
                ] = True

                rows[fill_index][
                    "depth_source"
                ] = "interpolated"


# ============================================================
# 相機座標
# ============================================================

def camera_xyz(
    x_px: float,
    y_px: float,
    depth_mm: float,
    image_width: int,
    image_height: int,
    args: argparse.Namespace,
) -> tuple[float, float, float]:
    scale_x = (
        image_width
        / args.intrinsic_width
    )

    scale_y = (
        image_height
        / args.intrinsic_height
    )

    fx = args.fx * scale_x
    fy = args.fy * scale_y
    cx = args.cx * scale_x
    cy = args.cy * scale_y

    x_mm = (
        (x_px - cx)
        * depth_mm
        / fx
    )

    y_mm = (
        (y_px - cy)
        * depth_mm
        / fy
    )

    return (
        round(x_mm, 3),
        round(y_mm, 3),
        round(depth_mm, 3),
    )


def distance_3d(
    first: tuple[float, float, float],
    second: tuple[float, float, float],
) -> float:
    return math.sqrt(
        sum(
            (b - a) ** 2
            for a, b in zip(
                first,
                second,
            )
        )
    )


def calculate_xyz(
    rows: list[dict[str, Any]],
    args: argparse.Namespace,
) -> None:
    for row in rows:
        row["camera_x_mm"] = None
        row["camera_y_mm"] = None
        row["camera_z_mm"] = None

        row["step_distance_mm"] = None
        row["outlier"] = False
        row["valid"] = False
        row["invalid_reason"] = ""
        row["quality"] = "D"

        row["coordinate_frame"] = (
            "camera_fixed_per_task"
        )

        row[
            "robot_execution_ready"
        ] = False

        if row["state_filtered"] == "NONE":
            row["invalid_reason"] = (
                "state_none"
            )
            continue

        x = row["x_px_filtered"]
        y = row["y_px_filtered"]

        depth = row[
            "depth_temporal_median_mm"
        ]

        if (
            not is_number(x)
            or not is_number(y)
        ):
            row["invalid_reason"] = (
                "missing_pixel"
            )
            continue

        if not is_number(depth):
            row["invalid_reason"] = (
                "missing_depth"
            )
            continue

        x_mm, y_mm, z_mm = camera_xyz(
            x_px=float(x),
            y_px=float(y),
            depth_mm=float(depth),
            image_width=row["image_width"],
            image_height=row["image_height"],
            args=args,
        )

        row["camera_x_mm"] = x_mm
        row["camera_y_mm"] = y_mm
        row["camera_z_mm"] = z_mm

        row["valid"] = True

        if row["interpolated"]:
            row["quality"] = "C"
        elif (
            row["depth_mapping"]
            == "same_size"
        ):
            row["quality"] = "A"
        else:
            row["quality"] = "B"


def mark_isolated_jumps(
    rows: list[dict[str, Any]],
    args: argparse.Namespace,
) -> None:
    valid_indices = [
        index
        for index, row in enumerate(rows)
        if row["valid"]
    ]

    for position in range(
        1,
        len(valid_indices) - 1,
    ):
        previous = rows[
            valid_indices[position - 1]
        ]

        current = rows[
            valid_indices[position]
        ]

        following = rows[
            valid_indices[position + 1]
        ]

        previous_xyz = (
            previous["camera_x_mm"],
            previous["camera_y_mm"],
            previous["camera_z_mm"],
        )

        current_xyz = (
            current["camera_x_mm"],
            current["camera_y_mm"],
            current["camera_z_mm"],
        )

        next_xyz = (
            following["camera_x_mm"],
            following["camera_y_mm"],
            following["camera_z_mm"],
        )

        first_jump = distance_3d(
            previous_xyz,
            current_xyz,
        )

        second_jump = distance_3d(
            current_xyz,
            next_xyz,
        )

        neighbour_distance = distance_3d(
            previous_xyz,
            next_xyz,
        )

        isolated = (
            first_jump
            > args.max_camera_step_mm
            and second_jump
            > args.max_camera_step_mm
            and neighbour_distance
            <= args.max_camera_step_mm
        )

        if isolated:
            current["outlier"] = True
            current["valid"] = False
            current["quality"] = "D"

            current["invalid_reason"] = (
                "isolated_camera_jump"
            )


def calculate_steps(
    rows: list[dict[str, Any]],
) -> None:
    previous = None

    for row in rows:
        if not row["valid"]:
            continue

        current_xyz = (
            row["camera_x_mm"],
            row["camera_y_mm"],
            row["camera_z_mm"],
        )

        if previous is not None:
            row["step_distance_mm"] = round(
                distance_3d(
                    previous,
                    current_xyz,
                ),
                3,
            )

        previous = current_xyz


def assign_segments(
    rows: list[dict[str, Any]],
) -> None:
    segment_id = 0
    previous_type = "none"

    for row in rows:
        if not row["valid"]:
            segment_type = "none"

        elif row["state_filtered"] == "DOWN":
            segment_type = "welding"

        elif row["state_filtered"] == "UP":
            segment_type = "travel"

        else:
            segment_type = "none"

        if (
            segment_type != "none"
            and segment_type != previous_type
        ):
            segment_id += 1

        row["segment_type"] = segment_type

        row["segment_id"] = (
            segment_id
            if segment_type != "none"
            else None
        )

        previous_type = segment_type


# ============================================================
# 匯出 CSV、Excel、預覽圖
# ============================================================

def write_csv(
    path: Path,
    rows: list[dict[str, Any]],
) -> None:
    with path.open(
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=TRAJECTORY_FIELDS,
            extrasaction="ignore",
        )

        writer.writeheader()
        writer.writerows(rows)


def make_summary(
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    total = len(rows)

    detected = sum(
        row["state_raw"] != "NONE"
        for row in rows
    )

    local_depth = sum(
        row["depth_local_median_mm"]
        is not None
        for row in rows
    )

    valid = sum(
        bool(row["valid"])
        for row in rows
    )

    interpolated = sum(
        bool(row["interpolated"])
        for row in rows
    )

    outliers = sum(
        bool(row["outlier"])
        for row in rows
    )

    mapping_counts = Counter(
        row["depth_mapping"]
        for row in rows
    )

    return {
        "created_at": datetime.now().isoformat(
            timespec="seconds"
        ),
        "total_samples": total,
        "detected_samples": detected,
        "detection_rate": (
            detected / total if total else 0
        ),
        "local_depth_samples": local_depth,
        "local_depth_rate": (
            local_depth / total
            if total
            else 0
        ),
        "valid_trajectory_samples": valid,
        "valid_trajectory_rate": (
            valid / total if total else 0
        ),
        "interpolated_samples": interpolated,
        "outlier_samples": outliers,
        "depth_mapping_counts": dict(
            mapping_counts
        ),
        "coordinate_frame": (
            "camera_fixed_per_task"
        ),
        "camera_pose_fixed_within_task": True,
        "cross_task_coordinates_comparable": False,
        "robot_execution_ready": False,
    }


def write_excel(
    path: Path,
    rows: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font,
            PatternFill,
        )
        from openpyxl.utils import (
            get_column_letter,
        )
    except ImportError as exc:
        raise SystemExit(
            "缺少 openpyxl，請執行：\n"
            "pip install openpyxl"
        ) from exc

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Trajectory"

    sheet.append(TRAJECTORY_FIELDS)

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    invalid_fill = PatternFill(
        "solid",
        fgColor="F4CCCC",
    )

    interpolated_fill = PatternFill(
        "solid",
        fgColor="FFF2CC",
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        sheet.append([
            row.get(field)
            for field in TRAJECTORY_FIELDS
        ])

        row_number = sheet.max_row

        if not row["valid"]:
            for cell in sheet[row_number]:
                cell.fill = invalid_fill

        elif row["interpolated"]:
            for cell in sheet[row_number]:
                cell.fill = interpolated_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, field in enumerate(
        TRAJECTORY_FIELDS,
        start=1,
    ):
        sheet.column_dimensions[
            get_column_letter(column_index)
        ].width = min(
            max(len(field) + 2, 12),
            28,
        )

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    summary_sheet.append([
        "項目",
        "數值",
    ])

    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for key, value in summary.items():
        if isinstance(value, dict):
            value = json.dumps(
                value,
                ensure_ascii=False,
            )

        summary_sheet.append([
            key,
            value,
        ])

    summary_sheet.column_dimensions[
        "A"
    ].width = 36

    summary_sheet.column_dimensions[
        "B"
    ].width = 70

    workbook.save(path)


def create_preview(
    path: Path,
    rows: list[dict[str, Any]],
) -> None:
    if not rows:
        return

    background_path = Path(
        rows[0]["source"]
    )

    image = cv2.imread(
        str(background_path),
        cv2.IMREAD_COLOR,
    )

    if image is None:
        return

    segments: dict[
        int,
        list[tuple[int, int]],
    ] = {}

    for row in rows:
        if (
            not row["valid"]
            or row["segment_type"]
            != "welding"
            or row["segment_id"] is None
        ):
            continue

        point = (
            int(round(
                row["x_px_filtered"]
            )),
            int(round(
                row["y_px_filtered"]
            )),
        )

        segments.setdefault(
            int(row["segment_id"]),
            [],
        ).append(point)

    for points in segments.values():
        if len(points) < 2:
            continue

        points_array = np.array(
            points,
            dtype=np.int32,
        ).reshape((-1, 1, 2))

        cv2.polylines(
            image,
            [points_array],
            False,
            (0, 0, 255),
            2,
            cv2.LINE_AA,
        )

    cv2.imwrite(
        str(path),
        image,
    )


# ============================================================
# 主程式
# ============================================================

def main() -> int:
    args = parse_args()

    predictions_path = resolve_path(
        args.predictions
    )

    if not predictions_path.is_file():
        raise SystemExit(
            f"找不到 predictions.csv："
            f"{predictions_path}"
        )

    if args.output_dir is None:
        output_dir = (
            predictions_path.parent
            / "trajectory"
        )
    else:
        output_dir = resolve_path(
            args.output_dir
        )

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    rows = read_predictions(
        predictions_path
    )

    attach_depth(
        rows,
        args,
    )

    temporal_filter(
        rows,
        args,
    )

    interpolate_short_gaps(
        rows,
        args,
    )

    calculate_xyz(
        rows,
        args,
    )

    mark_isolated_jumps(
        rows,
        args,
    )

    calculate_steps(rows)
    assign_segments(rows)

    summary = make_summary(rows)

    csv_path = (
        output_dir
        / "trajectory.csv"
    )

    excel_path = (
        output_dir
        / "trajectory.xlsx"
    )

    summary_path = (
        output_dir
        / "trajectory_summary.json"
    )

    preview_path = (
        output_dir
        / "trajectory_preview.png"
    )

    write_csv(
        csv_path,
        rows,
    )

    write_excel(
        excel_path,
        rows,
        summary,
    )

    summary_path.write_text(
        json.dumps(
            summary,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    create_preview(
        preview_path,
        rows,
    )

    print("\n軌跡匯出完成")
    print(f"CSV：{csv_path}")
    print(f"Excel：{excel_path}")
    print(f"Summary：{summary_path}")
    print(f"Preview：{preview_path}")

    print(
        "有效點："
        f"{summary['valid_trajectory_samples']}"
        f"/{summary['total_samples']} "
        f"({summary['valid_trajectory_rate']:.2%})"
    )

    print(
        "\n座標系：camera_fixed_per_task"
    )

    print(
        "不同影片不可直接拼接，"
        "目前不可直接送入手臂。"
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
